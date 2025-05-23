#!/usr/bin/env python3
import os
import json
from datetime import datetime, timedelta
from pathlib import Path
from typing import Dict, List, Tuple, Optional
from collections import defaultdict

import pandas as pd
import typer
from git import Repo
from git.objects.commit import Commit
from rich.console import Console
from rich.table import Table
from dateutil.relativedelta import relativedelta
from openpyxl.utils.dataframe import dataframe_to_rows

app = typer.Typer()
console = Console()

class GitAnalyzer:
    def __init__(self, repo_path: str):
        self.repo_path = Path(repo_path).resolve()
        self.repo = Repo(str(self.repo_path))
        self.reports_dir = Path("relatorios")
        self.reports_dir.mkdir(exist_ok=True)
        self.power_bi_dir = Path("power_bi_data")
        self.power_bi_dir.mkdir(exist_ok=True)

    def get_recent_developer(self) -> Tuple[str, str]:
        """Obtém o desenvolvedor mais recente que fez commit no repositório."""
        recent_commit = self.repo.head.commit
        return recent_commit.author.name, recent_commit.author.email

    def get_all_branches(self) -> List[str]:
        """Retorna lista de todas as branches do repositório."""
        return [ref.name for ref in self.repo.references if not ref.name.startswith('origin/')]

    def classify_branch(self, branch_name: str) -> str:
        """Classifica a branch como PRD ou HML."""
        if branch_name.lower() in ['main', 'master']:
            return 'PRD'
        return 'HML'

    def get_most_recent_name_for_email(self, commits_data: List[Dict]) -> Dict[str, str]:
        """Retorna o nome mais recente usado para cada email."""
        email_to_name = {}
        email_to_date = {}
        
        for commit in commits_data:
            email = commit['email_autor']
            date = commit['data']
            name = commit['nome_autor']
            
            if email not in email_to_date or date > email_to_date[email]:
                email_to_date[email] = date
                email_to_name[email] = name
        
        return email_to_name

    def analyze_commits(self, author_emails: List[str] = None) -> Tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame, pd.DataFrame, pd.DataFrame, pd.DataFrame, pd.DataFrame]:
        """Analisa commits e retorna estatísticas."""
        commits_data = []
        branch_commits = defaultdict(list)
        
        # Analisa commits em cada branch
        for branch in self.get_all_branches():
            try:
                ambiente = self.classify_branch(branch)
                for commit in self.repo.iter_commits(branch):
                    # Se author_emails está definido, verifica se o email do commit está na lista
                    if author_emails and commit.author.email not in author_emails:
                        continue
                    
                    stats = commit.stats.total
                    commit_date = pd.to_datetime(commit.committed_datetime).tz_convert('UTC').tz_localize(None)
                    
                    commit_info = {
                        'data': commit_date,
                        'nome_autor': commit.author.name,
                        'email_autor': commit.author.email,
                        'branch': branch,
                        'ambiente': ambiente,
                        'arquivos_alterados': stats['files'],
                        'linhas_adicionadas': stats['insertions'],
                        'linhas_removidas': stats['deletions'],
                        'total_linhas': stats['insertions'] + stats['deletions'],
                        'commit_hash': commit.hexsha[:8]  # Primeiros 8 caracteres do hash
                    }
                    
                    commits_data.append(commit_info)
                    branch_commits[branch].append(commit_info)
            except:
                console.print(f"[yellow]Aviso: Não foi possível analisar a branch {branch}[/yellow]")
        
        if not commits_data:
            return pd.DataFrame(), pd.DataFrame(), pd.DataFrame(), pd.DataFrame(), pd.DataFrame(), pd.DataFrame(), pd.DataFrame()

        # Obtém o nome mais recente para cada email
        email_to_name = self.get_most_recent_name_for_email(commits_data)
        
        # Atualiza os nomes dos autores para usar o mais recente
        for commit in commits_data:
            commit['nome_autor'] = email_to_name[commit['email_autor']]
            
        df = pd.DataFrame(commits_data)
        
        # Estatísticas diárias
        daily_stats = df.groupby([df['data'].dt.date, 'email_autor', 'branch', 'ambiente']).agg({
            'nome_autor': 'first',
            'arquivos_alterados': 'sum',
            'linhas_adicionadas': 'sum',
            'linhas_removidas': 'sum',
            'total_linhas': 'sum',
            'commit_hash': 'count'  # Conta número de commits
        }).reset_index()
        daily_stats.rename(columns={'commit_hash': 'total_commits'}, inplace=True)
        
        # Estatísticas mensais
        df['mes'] = df['data'].dt.to_period('M')
        monthly_stats = df.groupby(['mes', 'email_autor', 'branch', 'ambiente']).agg({
            'nome_autor': 'first',
            'arquivos_alterados': 'sum',
            'linhas_adicionadas': 'sum',
            'linhas_removidas': 'sum',
            'total_linhas': 'sum',
            'commit_hash': 'count'  # Conta número de commits
        }).reset_index()
        monthly_stats.rename(columns={'commit_hash': 'total_commits'}, inplace=True)
        
        # Totais diários
        daily_totals = df.groupby([df['data'].dt.date, 'ambiente']).agg({
            'arquivos_alterados': 'sum',
            'linhas_adicionadas': 'sum',
            'linhas_removidas': 'sum',
            'total_linhas': 'sum',
            'commit_hash': 'count'  # Conta número de commits
        }).reset_index()
        daily_totals.rename(columns={'commit_hash': 'total_commits'}, inplace=True)
        daily_totals = daily_totals.sort_values('data', ascending=False)
        
        # Totais mensais
        monthly_totals = df.groupby(['mes', 'ambiente']).agg({
            'arquivos_alterados': 'sum',
            'linhas_adicionadas': 'sum',
            'linhas_removidas': 'sum',
            'total_linhas': 'sum',
            'commit_hash': 'count'  # Conta número de commits
        }).reset_index()
        monthly_totals.rename(columns={'commit_hash': 'total_commits'}, inplace=True)
        monthly_totals = monthly_totals.sort_values('mes', ascending=False)
        
        # Estatísticas por branch
        branch_stats = df.groupby(['branch', 'ambiente', 'email_autor']).agg({
            'nome_autor': 'first',
            'arquivos_alterados': 'sum',
            'linhas_adicionadas': 'sum',
            'linhas_removidas': 'sum',
            'total_linhas': 'sum',
            'commit_hash': 'count'  # Conta número de commits
        }).reset_index()
        branch_stats.rename(columns={'commit_hash': 'total_commits'}, inplace=True)
        
        # Resumo por ambiente
        env_stats = df.groupby(['ambiente']).agg({
            'arquivos_alterados': 'sum',
            'linhas_adicionadas': 'sum',
            'linhas_removidas': 'sum',
            'total_linhas': 'sum',
            'commit_hash': 'count',  # Conta número de commits
            'branch': 'nunique'  # Conta número único de branches
        }).reset_index()
        env_stats.rename(columns={
            'commit_hash': 'total_commits',
            'branch': 'total_branches'
        }, inplace=True)
        
        # Resumo geral
        summary_stats = df.groupby(['email_autor', 'ambiente']).agg({
            'nome_autor': 'first',
            'arquivos_alterados': 'sum',
            'linhas_adicionadas': 'sum',
            'linhas_removidas': 'sum',
            'total_linhas': 'sum',
            'commit_hash': 'count',  # Conta número de commits
            'branch': 'nunique'  # Conta número único de branches
        }).reset_index()
        summary_stats.rename(columns={
            'commit_hash': 'total_commits',
            'branch': 'total_branches'
        }, inplace=True)
        
        # Converte o período mensal para string para melhor visualização no Excel
        monthly_stats['mes'] = monthly_stats['mes'].astype(str)
        monthly_totals['mes'] = monthly_totals['mes'].astype(str)
        
        return daily_stats, monthly_stats, branch_stats, summary_stats, daily_totals, monthly_totals, env_stats

    def save_reports(self, daily_stats: pd.DataFrame, monthly_stats: pd.DataFrame, 
                    branch_stats: pd.DataFrame, summary_stats: pd.DataFrame,
                    daily_totals: pd.DataFrame, monthly_totals: pd.DataFrame,
                    env_stats: pd.DataFrame, author_name: str, author_email: str) -> str:
        """Salva relatórios em arquivo Excel."""
        try:
            data_hora = datetime.now().strftime("%d-%m-%Y-%H-%M")
            nome_autor_limpo = "".join(c for c in author_name.lower() if c.isalnum() or c == '_')
            nome_arquivo = f"relatorio_git_{data_hora}_{nome_autor_limpo}.xlsx"
            caminho_arquivo = self.reports_dir / nome_arquivo

            self.reports_dir.mkdir(exist_ok=True)

            with pd.ExcelWriter(str(caminho_arquivo), engine='openpyxl') as writer:
                # Resumo Geral - Por Autor e Ambiente
                workbook = writer.book
                resumo_sheet = workbook.create_sheet('Resumo Geral')
                writer.sheets['Resumo Geral'] = resumo_sheet

                # Estilo para títulos das seções
                from openpyxl.styles import PatternFill, Font, Alignment
                titulo_font = Font(bold=True, size=12)
                
                # Cores para cada seção
                cores_secoes = {
                    'autor': 'E6F3FF',  # Azul claro
                    'ambiente': 'E6FFE6',  # Verde claro
                    'diario': 'FFE6E6',  # Vermelho claro
                    'mensal': 'FFF3E6'   # Laranja claro
                }

                # Seção 1: Resumo por Autor
                resumo_sheet.merge_cells('A1:H1')
                titulo_cell = resumo_sheet['A1']
                titulo_cell.value = "RESUMO POR AUTOR E AMBIENTE"
                titulo_cell.font = titulo_font
                titulo_cell.alignment = Alignment(horizontal='center')
                titulo_cell.fill = PatternFill(start_color=cores_secoes['autor'], end_color=cores_secoes['autor'], fill_type='solid')

                # Adiciona os dados do resumo
                start_row = 2
                for r_idx, row in enumerate(dataframe_to_rows(summary_stats, index=False, header=True), start_row):
                    for c_idx, value in enumerate(row, 1):
                        cell = resumo_sheet.cell(row=r_idx, column=c_idx, value=value)
                        if r_idx == start_row:  # Cabeçalho
                            cell.font = Font(bold=True)
                        cell.fill = PatternFill(start_color=cores_secoes['autor'], end_color=cores_secoes['autor'], fill_type='solid')

                # Seção 2: Resumo por Ambiente
                start_row = len(summary_stats) + 4
                resumo_sheet.merge_cells(f'A{start_row}:H{start_row}')
                titulo_cell = resumo_sheet[f'A{start_row}']
                titulo_cell.value = "RESUMO POR AMBIENTE (PRD/HML)"
                titulo_cell.font = titulo_font
                titulo_cell.alignment = Alignment(horizontal='center')
                titulo_cell.fill = PatternFill(start_color=cores_secoes['ambiente'], end_color=cores_secoes['ambiente'], fill_type='solid')

                # Adiciona os dados do ambiente
                start_row += 1
                for r_idx, row in enumerate(dataframe_to_rows(env_stats, index=False, header=True), start_row):
                    for c_idx, value in enumerate(row, 1):
                        cell = resumo_sheet.cell(row=r_idx, column=c_idx, value=value)
                        if r_idx == start_row:  # Cabeçalho
                            cell.font = Font(bold=True)
                        cell.fill = PatternFill(start_color=cores_secoes['ambiente'], end_color=cores_secoes['ambiente'], fill_type='solid')

                # Seção 3: Totais Diários
                start_row = len(summary_stats) + len(env_stats) + 7
                resumo_sheet.merge_cells(f'A{start_row}:H{start_row}')
                titulo_cell = resumo_sheet[f'A{start_row}']
                titulo_cell.value = "TOTAIS DIÁRIOS"
                titulo_cell.font = titulo_font
                titulo_cell.alignment = Alignment(horizontal='center')
                titulo_cell.fill = PatternFill(start_color=cores_secoes['diario'], end_color=cores_secoes['diario'], fill_type='solid')

                # Adiciona os dados diários
                start_row += 1
                for r_idx, row in enumerate(dataframe_to_rows(daily_totals, index=False, header=True), start_row):
                    for c_idx, value in enumerate(row, 1):
                        cell = resumo_sheet.cell(row=r_idx, column=c_idx, value=value)
                        if r_idx == start_row:  # Cabeçalho
                            cell.font = Font(bold=True)
                        cell.fill = PatternFill(start_color=cores_secoes['diario'], end_color=cores_secoes['diario'], fill_type='solid')

                # Seção 4: Totais Mensais
                start_row = len(summary_stats) + len(env_stats) + len(daily_totals) + 10
                resumo_sheet.merge_cells(f'A{start_row}:H{start_row}')
                titulo_cell = resumo_sheet[f'A{start_row}']
                titulo_cell.value = "TOTAIS MENSAIS"
                titulo_cell.font = titulo_font
                titulo_cell.alignment = Alignment(horizontal='center')
                titulo_cell.fill = PatternFill(start_color=cores_secoes['mensal'], end_color=cores_secoes['mensal'], fill_type='solid')

                # Adiciona os dados mensais
                start_row += 1
                for r_idx, row in enumerate(dataframe_to_rows(monthly_totals, index=False, header=True), start_row):
                    for c_idx, value in enumerate(row, 1):
                        cell = resumo_sheet.cell(row=r_idx, column=c_idx, value=value)
                        if r_idx == start_row:  # Cabeçalho
                            cell.font = Font(bold=True)
                        cell.fill = PatternFill(start_color=cores_secoes['mensal'], end_color=cores_secoes['mensal'], fill_type='solid')

                # Outras abas
                branch_stats.to_excel(writer, sheet_name='Estatísticas por Branch', index=False)
                daily_stats.to_excel(writer, sheet_name='Estatísticas Diárias', index=False)
                monthly_stats.to_excel(writer, sheet_name='Estatísticas Mensais', index=False)

                # Aba de informações
                info_df = pd.DataFrame([{
                    'Data do Relatório': datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                    'Autor': author_name,
                    'Email': author_email,
                    'Repositório': str(self.repo_path)
                }])
                info_df.to_excel(writer, sheet_name='Informações', index=False)

                # Ajusta as larguras das colunas e adiciona filtros
                for sheet_name in writer.sheets:
                    worksheet = writer.sheets[sheet_name]
                    
                    if sheet_name != 'Informações':
                        if sheet_name == 'Resumo Geral':
                            # Filtros para cada seção do resumo
                            sections = [
                                (2, len(summary_stats) + 1),
                                (len(summary_stats) + 5, len(summary_stats) + 4 + len(env_stats)),
                                (len(summary_stats) + len(env_stats) + 8, len(summary_stats) + len(env_stats) + 7 + len(daily_totals)),
                                (len(summary_stats) + len(env_stats) + len(daily_totals) + 11, len(summary_stats) + len(env_stats) + len(daily_totals) + 10 + len(monthly_totals))
                            ]
                            
                            for start_row, end_row in sections:
                                if end_row > start_row:
                                    worksheet.auto_filter.ref = f"A{start_row}:H{end_row}"
                        else:
                            # Filtro para toda a aba
                            max_row = worksheet.max_row
                            if max_row > 1:
                                worksheet.auto_filter.ref = worksheet.dimensions
                    
                    # Ajusta larguras das colunas
                    for idx, col in enumerate(worksheet.columns, 1):
                        max_length = 0
                        column = worksheet.column_dimensions[chr(64 + idx)]
                        
                        for cell in col:
                            try:
                                if len(str(cell.value)) > max_length:
                                    max_length = len(str(cell.value))
                            except:
                                pass
                        
                        adjusted_width = (max_length + 2)
                        column.width = adjusted_width

            return nome_arquivo

        except Exception as e:
            console.print(f"[red]Erro ao salvar relatório Excel: {str(e)}[/red]")
            raise

    def display_report(self, daily_stats: pd.DataFrame, monthly_stats: pd.DataFrame, 
                      branch_stats: pd.DataFrame, summary_stats: pd.DataFrame,
                      daily_totals: pd.DataFrame, monthly_totals: pd.DataFrame,
                      env_stats: pd.DataFrame, author_name: str, author_email: str):
        """Exibe um relatório formatado no console."""
        console.print(f"\n[bold blue]Relatório de Métricas Git para {author_name} ({author_email})[/bold blue]")
        
        # Tabela de resumo geral
        summary_table = Table(title="\nResumo Geral - Por Autor e Ambiente")
        summary_table.add_column("Autor", justify="left")
        summary_table.add_column("Email", justify="left")
        summary_table.add_column("Ambiente", justify="left")
        summary_table.add_column("Total de Branches", justify="right")
        summary_table.add_column("Arquivos Alterados", justify="right")
        summary_table.add_column("Linhas Adicionadas", justify="right")
        summary_table.add_column("Linhas Removidas", justify="right")
        summary_table.add_column("Total de Linhas", justify="right")
        
        for _, row in summary_stats.iterrows():
            summary_table.add_row(
                str(row['nome_autor']),
                str(row['email_autor']),
                str(row['ambiente']),
                str(row['total_branches']),
                str(row['arquivos_alterados']),
                str(row['linhas_adicionadas']),
                str(row['linhas_removidas']),
                str(row['total_linhas'])
            )
        
        console.print(summary_table)
        
        # Tabela de totais diários
        daily_totals_table = Table(title="\nTotais Diários")
        daily_totals_table.add_column("Data", justify="left")
        daily_totals_table.add_column("Ambiente", justify="left")
        daily_totals_table.add_column("Arquivos Alterados", justify="right")
        daily_totals_table.add_column("Linhas Adicionadas", justify="right")
        daily_totals_table.add_column("Linhas Removidas", justify="right")
        daily_totals_table.add_column("Total de Linhas", justify="right")
        
        for _, row in daily_totals.head(7).iterrows():  # Mostra apenas os últimos 7 dias
            daily_totals_table.add_row(
                str(row['data']),
                str(row['ambiente']),
                str(row['arquivos_alterados']),
                str(row['linhas_adicionadas']),
                str(row['linhas_removidas']),
                str(row['total_linhas'])
            )
        
        console.print(daily_totals_table)
        
        # Tabela de totais mensais
        monthly_totals_table = Table(title="\nTotais Mensais")
        monthly_totals_table.add_column("Mês", justify="left")
        monthly_totals_table.add_column("Ambiente", justify="left")
        monthly_totals_table.add_column("Arquivos Alterados", justify="right")
        monthly_totals_table.add_column("Linhas Adicionadas", justify="right")
        monthly_totals_table.add_column("Linhas Removidas", justify="right")
        monthly_totals_table.add_column("Total de Linhas", justify="right")
        
        for _, row in monthly_totals.head(3).iterrows():  # Mostra apenas os últimos 3 meses
            monthly_totals_table.add_row(
                str(row['mes']),
                str(row['ambiente']),
                str(row['arquivos_alterados']),
                str(row['linhas_adicionadas']),
                str(row['linhas_removidas']),
                str(row['total_linhas'])
            )
        
        console.print(monthly_totals_table)
        
        # Tabela de estatísticas por branch
        branch_table = Table(title="\nEstatísticas por Branch")
        branch_table.add_column("Branch", justify="left")
        branch_table.add_column("Ambiente", justify="left")
        branch_table.add_column("Autor", justify="left")
        branch_table.add_column("Email", justify="left")
        branch_table.add_column("Arquivos Alterados", justify="right")
        branch_table.add_column("Linhas Adicionadas", justify="right")
        branch_table.add_column("Linhas Removidas", justify="right")
        branch_table.add_column("Total de Linhas", justify="right")
        
        for _, row in branch_stats.iterrows():
            branch_table.add_row(
                str(row['branch']),
                str(row['ambiente']),
                str(row['nome_autor']),
                str(row['email_autor']),
                str(row['arquivos_alterados']),
                str(row['linhas_adicionadas']),
                str(row['linhas_removidas']),
                str(row['total_linhas'])
            )
        
        console.print(branch_table)
        
        # Tabela de estatísticas diárias
        daily_table = Table(title="\nEstatísticas Diárias (Últimos 7 dias)")
        daily_table.add_column("Data", justify="left")
        daily_table.add_column("Ambiente", justify="left")
        daily_table.add_column("Autor", justify="left")
        daily_table.add_column("Email", justify="left")
        daily_table.add_column("Arquivos Alterados", justify="right")
        daily_table.add_column("Linhas Adicionadas", justify="right")
        daily_table.add_column("Linhas Removidas", justify="right")
        daily_table.add_column("Total de Linhas", justify="right")
        
        recent_daily = daily_stats.sort_values('data', ascending=False).head(7)
        for _, row in recent_daily.iterrows():
            daily_table.add_row(
                str(row['data']),
                str(row['ambiente']),
                str(row['nome_autor']),
                str(row['email_autor']),
                str(row['arquivos_alterados']),
                str(row['linhas_adicionadas']),
                str(row['linhas_removidas']),
                str(row['total_linhas'])
            )
        
        console.print(daily_table)
        
        # Tabela de estatísticas mensais
        monthly_table = Table(title="\nEstatísticas Mensais (Últimos 3 meses)")
        monthly_table.add_column("Mês", justify="left")
        monthly_table.add_column("Ambiente", justify="left")
        monthly_table.add_column("Autor", justify="left")
        monthly_table.add_column("Email", justify="left")
        monthly_table.add_column("Arquivos Alterados", justify="right")
        monthly_table.add_column("Linhas Adicionadas", justify="right")
        monthly_table.add_column("Linhas Removidas", justify="right")
        monthly_table.add_column("Total de Linhas", justify="right")
        
        recent_monthly = monthly_stats.sort_values('mes', ascending=False).head(3)
        for _, row in recent_monthly.iterrows():
            monthly_table.add_row(
                str(row['mes']),
                str(row['ambiente']),
                str(row['nome_autor']),
                str(row['email_autor']),
                str(row['arquivos_alterados']),
                str(row['linhas_adicionadas']),
                str(row['linhas_removidas']),
                str(row['total_linhas'])
            )
        
        console.print(monthly_table)

    def generate_power_bi_data(self, author_emails: List[str] = None) -> str:
        """Gera arquivo de dados para integração com Power BI."""
        try:
            daily_stats, monthly_stats, branch_stats, summary_stats, daily_totals, monthly_totals, env_stats = self.analyze_commits(author_emails)
            
            if daily_stats.empty:
                raise ValueError("Nenhum dado encontrado para gerar o relatório")

            # Prepara dados para o Power BI com formatação adequada
            def prepare_dataframe(df):
                if df is None or df.empty:
                    return []
                df_copy = df.copy()
                if 'mes' in df_copy.columns:
                    df_copy['mes'] = df_copy['mes'].astype(str)
                if 'data' in df_copy.columns:
                    df_copy['data'] = pd.to_datetime(df_copy['data']).dt.strftime('%Y-%m-%d')
                return df_copy.to_dict(orient='records')

            power_bi_data = {
                "metadata": {
                    "ultima_atualizacao": datetime.now().isoformat(),
                    "repositorio": str(self.repo_path),
                    "versao": "1.0"
                },
                "dados": {
                    "metricas_diarias": prepare_dataframe(daily_stats),
                    "metricas_mensais": prepare_dataframe(monthly_stats),
                    "metricas_branch": prepare_dataframe(branch_stats),
                    "resumo_geral": prepare_dataframe(summary_stats),
                    "totais_diarios": prepare_dataframe(daily_totals),
                    "totais_mensais": prepare_dataframe(monthly_totals),
                    "metricas_ambiente": prepare_dataframe(env_stats)
                }
            }
            
            # Formata a data e hora no padrão solicitado
            data_hora = datetime.now().strftime("%d-%m-%Y-%H-%M")
            # Obtém o nome do autor do primeiro commit encontrado
            author_name = "desconhecido"
            for commit in self.repo.iter_commits():
                if commit.author.email in author_emails:
                    author_name = commit.author.name
                    break
            
            # Limpa o nome do autor para usar no nome do arquivo (remove caracteres especiais)
            nome_autor_limpo = "".join(c for c in author_name.lower() if c.isalnum() or c == '_')
            
            # Verifica se o diretório existe, se não, cria
            self.power_bi_dir.mkdir(exist_ok=True)
            
            # Salva o arquivo JSON principal
            json_file = self.power_bi_dir / f"git_metrics_{data_hora}_{nome_autor_limpo}.json"
            with open(str(json_file), 'w', encoding='utf-8') as f:
                json.dump(power_bi_data, f, ensure_ascii=False, indent=2)
            
            # Gera arquivo de configuração para o Power BI
            config = {
                "configuracao": {
                    "fonte_dados": str(json_file),
                    "frequencia_atualizacao": "daily",
                    "versao_metricas": "1.0",
                    "caminho_repositorio": str(self.repo_path),
                    "ultima_atualizacao": datetime.now().isoformat()
                },
                "metricas_disponiveis": [
                    "commits_por_dia",
                    "commits_por_ambiente",
                    "linhas_alteradas",
                    "atividade_por_autor",
                    "distribuicao_commits"
                ],
                "ambientes": {
                    "PRD": ["main", "master"],
                    "HML": "outros"
                }
            }
            
            # Salva o arquivo de configuração
            config_file = self.power_bi_dir / f"power_bi_config_{data_hora}_{nome_autor_limpo}.json"
            with open(str(config_file), 'w', encoding='utf-8') as f:
                json.dump(config, f, ensure_ascii=False, indent=2)

            return str(json_file)

        except Exception as e:
            console.print(f"[red]Erro ao gerar dados do Power BI: {str(e)}[/red]")
            raise

    def generate_power_bi_readme(self):
        """Gera arquivo README com instruções para o Power BI."""
        readme_content = """# Análise de Métricas Git com Power BI

## Estrutura dos Arquivos
- `git_metrics_*.json`: Arquivo principal com os dados das métricas
- `power_bi_config.json`: Arquivo de configuração do painel
- `medidas_dax.txt`: Modelo de medidas DAX para análises

## Como Utilizar
1. Abra o Power BI Desktop
2. Importe o arquivo JSON mais recente:
   - Clique em 'Obter Dados'
   - Selecione 'JSON'
   - Navegue até a pasta 'power_bi_data'
   - Escolha o arquivo mais recente
3. Copie e cole as medidas DAX do arquivo template
4. Monte as visualizações conforme instruções abaixo

## Estrutura do Painel

### 1. Página Inicial (Visão Geral)
- Cartões com métricas principais:
  * Total de Commits
  * Total de Linhas de Código
  * Número de Autores
  * Quantidade de Branches
- Gráfico de rosca: Distribuição PRD vs HML
- Linha do tempo de commits (gráfico de área)
- Principais contribuidores (gráfico de barras)

### 2. Análise Detalhada
- Métricas por branch (mapa de árvore)
- Evolução temporal (gráfico de linha)
- Mapa de calor de atividades
- Filtros personalizados:
  * Por período
  * Por autor
  * Por ambiente (PRD/HML)

### 3. Produtividade
- Métricas por desenvolvedor
- Tendências de commits
- Análise comparativa
- Indicadores de desempenho:
  * Média de commits por dia
  * Linhas de código por commit
  * Atividade por ambiente

## Atualização de Dados
1. Execute o script Python para gerar novos dados
2. No Power BI Desktop:
   - Clique em 'Atualizar' para dados mais recentes
   - Configure atualização automática se necessário

## Medidas DAX Disponíveis
- Contagem total de commits
- Commits por ambiente (PRD/HML)
- Médias e tendências
- Métricas de produtividade
- Análises temporais

## Filtros Globais Disponíveis
1. Período:
   - Diário
   - Semanal
   - Mensal
2. Ambiente:
   - Produção (PRD)
   - Homologação (HML)
3. Autor:
   - Individual
   - Múltiplos autores
4. Branch:
   - Principal
   - Específica

## Dicas de Utilização
1. Use os filtros globais para análises específicas
2. Explore diferentes visualizações dos dados
3. Exporte relatórios quando necessário
4. Configure alertas para métricas importantes
5. Salve visualizações personalizadas

## Manutenção e Suporte
1. Documentação:
   - Consulte este guia para referência
   - Verifique atualizações periódicas
2. Logs:
   - Monitore os logs de execução
   - Verifique erros e avisos
3. Suporte:
   - Entre em contato com a equipe de desenvolvimento
   - Reporte problemas encontrados

## Boas Práticas
1. Atualize os dados regularmente
2. Mantenha backups das configurações personalizadas
3. Documente alterações realizadas
4. Compartilhe insights relevantes
5. Utilize os filtros adequadamente

## Resolução de Problemas
1. Dados não atualizados:
   - Verifique a execução do script
   - Confira a conexão com o repositório
2. Visualizações incorretas:
   - Revise as medidas DAX
   - Verifique os filtros aplicados
3. Erros de carregamento:
   - Confirme o formato do arquivo JSON
   - Verifique permissões de acesso

## Recursos Adicionais
- Documentação do projeto
- Guias de análise
- Exemplos de uso
- Modelos de relatório

Para mais informações ou suporte:
- Consulte a documentação completa do projeto
- Entre em contato com a equipe de desenvolvimento
- Verifique atualizações e novidades"""

        readme_file = self.power_bi_dir / "README.md"
        with open(readme_file, 'w', encoding='utf-8') as f:
            f.write(readme_content)

def validate_path(path: str) -> str:
    """Valida e normaliza o caminho do repositório."""
    try:
        repo_path = Path(path).resolve()
        if not repo_path.exists():
            raise typer.BadParameter(f"Caminho do repositório não existe: {path}")
        if not (repo_path / ".git").exists():
            raise typer.BadParameter(f"Não é um repositório git: {path}")
        return str(repo_path)
    except Exception as e:
        raise typer.BadParameter(str(e))

@app.command()
def analyze(
    repo_path: str = typer.Option(..., "--path", "-p", help="Caminho do repositório git para análise"),
    author_emails: Optional[List[str]] = typer.Option(None, "--author", "-a", help="Filtrar por email(s) do(s) autor(es). Pode ser especificado múltiplas vezes.")
):
    """Analisa um repositório git e gera relatórios de contribuição em Excel e Power BI."""
    try:
        repo_path = validate_path(repo_path)
        analyzer = GitAnalyzer(repo_path)
        
        if not author_emails:
            author_name, author_email = analyzer.get_recent_developer()
            author_emails = [author_email]
            console.print(f"[yellow]Nenhum autor especificado. Usando contribuidor mais recente: {author_name} ({author_email})[/yellow]\n")
        
        # Obtém os nomes dos autores
        author_names = {}
        for commit in analyzer.repo.iter_commits():
            if commit.author.email in author_emails and commit.author.email not in author_names:
                author_names[commit.author.email] = commit.author.name
        
        # Se algum email não foi encontrado, usa "Desconhecido"
        for email in author_emails:
            if email not in author_names:
                author_names[email] = "Desconhecido"
        
        daily_stats, monthly_stats, branch_stats, summary_stats, daily_totals, monthly_totals, env_stats = analyzer.analyze_commits(author_emails)
        
        if daily_stats.empty:
            console.print(f"[red]Nenhum commit encontrado para os autores especificados[/red]")
            return
        
        # Gera nome do arquivo com base no primeiro autor (ou usa 'multiple' se houver mais de um)
        if len(author_emails) == 1:
            author_name = author_names[author_emails[0]]
            author_email = author_emails[0]
        else:
            author_name = "multiple_authors"
            author_email = "multiple"
        
        # Gera relatório Excel
        excel_file = analyzer.save_reports(daily_stats, monthly_stats, branch_stats, summary_stats, 
                                       daily_totals, monthly_totals, env_stats, author_name, author_email)
        
        # Exibe relatório no console
        analyzer.display_report(daily_stats, monthly_stats, branch_stats, summary_stats, 
                            daily_totals, monthly_totals, env_stats, author_name, author_email)
        
        console.print(f"\n[green]Relatório Excel foi salvo como '{excel_file}' no diretório 'relatorios'.[/green]")
        
        # Gera arquivos para Power BI
        json_file = analyzer.generate_power_bi_data(author_emails)
        analyzer.generate_power_bi_readme()
        console.print(f"\n[green]Dados para Power BI gerados em '{json_file}'[/green]")
        console.print("[blue]Consulte o arquivo README.md no diretório 'power_bi_data' para instruções de uso.[/blue]")
        
        # Exibe resumo final
        console.print("\n[bold green]Geração de relatórios concluída:[/bold green]")
        console.print(f"1. Relatório Excel: {excel_file}")
        console.print(f"2. Dados Power BI: {json_file}")
        console.print("3. Documentação: power_bi_data/README.md")
        console.print("\n[bold blue]Próximos passos:[/bold blue]")
        console.print("1. Abra o relatório Excel para análise detalhada")
        console.print("2. Importe os dados JSON no Power BI Desktop")
        console.print("3. Siga as instruções no README.md para configurar o dashboard")
        
    except Exception as e:
        console.print(f"[red]Erro: {str(e)}[/red]")
        raise typer.Exit(1)

if __name__ == "__main__":
    app() 