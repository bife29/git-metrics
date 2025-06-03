from typing import List
from pathlib import Path
from domain.entities.author import Author
from infrastructure.repositories.git_repository import GitRepository
from application.commands.generate_excel_report_command import GenerateExcelReportCommand
from application.queries.get_commit_statistics_query import CommitStatisticsQuery
from datetime import datetime
import pandas as pd
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from application.queries.repository_metrics import RepositoryMetricsQuery

class AnalyzeRepositoryCommand:
    def __init__(self, repository: GitRepository):
        self.repository = repository
        self.output_dir = Path("reports")

    def execute(self, authors: List[Author]) -> str:
        """
        Executa a análise do repositório e gera o relatório Excel.
        Retorna o caminho do arquivo Excel gerado.
        """
        # Verifica se há dados para os autores
        has_data = False
        for author in authors:
            if self.repository.get_commits_by_author(author):
                has_data = True
                break
                
        if not has_data:
            raise ValueError("Nenhum dado encontrado para gerar o relatório")
        
        # Cria diretório de saída se não existir
        self.output_dir.mkdir(parents=True, exist_ok=True)
        
        # Gera nome do arquivo com timestamp
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        excel_path = self.output_dir / f"git_metrics_{timestamp}.xlsx"
        
        # Obtém métricas do repositório
        query = RepositoryMetricsQuery(self.repository)
        metrics = query.execute(authors)
        
        # Cria DataFrames para cada seção
        dfs = {
            'Resumo Geral': pd.DataFrame(metrics['resumo_autor_ambiente']),
            'Resumo por Ambiente': pd.DataFrame(metrics['resumo_ambiente']),
            'Totais Diários': pd.DataFrame(metrics['totais_diarios']),
            'Totais Mensais': pd.DataFrame(metrics['totais_mensais'])
        }
        
        # Cria o arquivo Excel
        with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
            for sheet_name, df in dfs.items():
                if not df.empty:
                    df.to_excel(writer, sheet_name=sheet_name, index=False)
                    self._format_sheet(writer.sheets[sheet_name])
            
            # Adiciona aba com informações dos autores
            authors_df = pd.DataFrame([
                {'Nome': author.name, 'Email': author.email}
                for author in authors
            ])
            authors_df.to_excel(writer, sheet_name='Autores', index=False)
            self._format_sheet(writer.sheets['Autores'])
        
        return str(excel_path)
    
    def _format_sheet(self, worksheet):
        """Aplica formatação padrão para uma aba."""
        # Cores
        header_fill = PatternFill(start_color='1F4E78', end_color='1F4E78', fill_type='solid')
        header_font = Font(color='FFFFFF', bold=True)
        
        # Borda
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Formata cabeçalho
        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border
        
        # Formata células de dados
        for row in worksheet.iter_rows(min_row=2):
            for cell in row:
                cell.border = border
                cell.alignment = Alignment(horizontal='center', vertical='center')
                
                # Aplica cores alternadas nas linhas
                if cell.row % 2 == 0:
                    cell.fill = PatternFill(start_color='F5F5F5', end_color='F5F5F5', fill_type='solid')
        
        # Ajusta larguras das colunas
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            adjusted_width = (max_length + 2)
            worksheet.column_dimensions[column_letter].width = min(adjusted_width, 30)
        
        # Adiciona filtros
        if worksheet.max_row > 1:
            worksheet.auto_filter.ref = worksheet.dimensions 