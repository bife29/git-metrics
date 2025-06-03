from dataclasses import dataclass
from pathlib import Path
from typing import List, Optional
from datetime import datetime

import pandas as pd
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

from domain.entities.author import Author
from application.queries.get_commit_statistics_query import CommitStatisticsQuery

@dataclass
class GenerateExcelReportCommand:
    """
    Command para gerar relatório Excel.
    Segue o padrão CQS (Command Query Separation).
    """
    statistics_query: CommitStatisticsQuery
    output_dir: Path
    authors: List[Author]
    
    def execute(self) -> str:
        """
        Executa o comando para gerar o relatório Excel.
        Retorna o caminho do arquivo Excel gerado.
        """
        # Cria o diretório de saída se não existir
        self.output_dir.mkdir(parents=True, exist_ok=True)
        
        # Gera o nome do arquivo com timestamp
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        excel_path = self.output_dir / f"git_metrics_{timestamp}.xlsx"
        
        # Obtém os dados das estatísticas
        daily_df, monthly_df, branch_df = self.statistics_query.execute()
        
        # Cria o arquivo Excel
        with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
            # Cria a aba de resumo geral primeiro
            workbook = writer.book
            resumo_sheet = workbook.create_sheet('Resumo Geral', 0)
            
            # Cores para cada seção
            cores_secoes = {
                'autor': 'E6F3FF',    # Azul claro
                'ambiente': 'E6FFE6',  # Verde claro
                'diario': 'FFE6E6',   # Vermelho claro
                'mensal': 'FFF3E6'    # Laranja claro
            }
            
            # Estilo para títulos das seções
            titulo_font = Font(bold=True, size=12)
            
            # 1. RESUMO POR AUTOR E AMBIENTE
            resumo_sheet.merge_cells('A1:H1')
            titulo_cell = resumo_sheet['A1']
            titulo_cell.value = "RESUMO POR AUTOR E AMBIENTE"
            titulo_cell.font = titulo_font
            titulo_cell.alignment = Alignment(horizontal='center')
            titulo_cell.fill = PatternFill(start_color=cores_secoes['autor'], end_color=cores_secoes['autor'], fill_type='solid')
            
            # Adiciona os dados do resumo por autor
            start_row = 2
            autor_df = branch_df.copy()
            for r_idx, row in enumerate(dataframe_to_rows(autor_df, index=False, header=True), start_row):
                for c_idx, value in enumerate(row, 1):
                    cell = resumo_sheet.cell(row=r_idx, column=c_idx, value=value)
                    if r_idx == start_row:  # Cabeçalho
                        cell.font = Font(bold=True)
                        cell.fill = PatternFill(start_color='1F4E78', end_color='1F4E78', fill_type='solid')
                        cell.font = Font(bold=True, color='FFFFFF')
                    else:
                        cell.fill = PatternFill(start_color=cores_secoes['autor'], end_color=cores_secoes['autor'], fill_type='solid')
                    cell.border = Border(
                        left=Side(style='thin'),
                        right=Side(style='thin'),
                        top=Side(style='thin'),
                        bottom=Side(style='thin')
                    )
                    cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # 2. RESUMO POR AMBIENTE
            start_row = len(autor_df) + 4
            resumo_sheet.merge_cells(f'A{start_row}:H{start_row}')
            titulo_cell = resumo_sheet[f'A{start_row}']
            titulo_cell.value = "RESUMO POR AMBIENTE (PRD/HML)"
            titulo_cell.font = titulo_font
            titulo_cell.alignment = Alignment(horizontal='center')
            titulo_cell.fill = PatternFill(start_color=cores_secoes['ambiente'], end_color=cores_secoes['ambiente'], fill_type='solid')
            
            # Adiciona os dados do ambiente
            start_row += 1
            env_df = branch_df.groupby('ambiente').sum().reset_index()
            for r_idx, row in enumerate(dataframe_to_rows(env_df, index=False, header=True), start_row):
                for c_idx, value in enumerate(row, 1):
                    cell = resumo_sheet.cell(row=r_idx, column=c_idx, value=value)
                    if r_idx == start_row:  # Cabeçalho
                        cell.font = Font(bold=True)
                        cell.fill = PatternFill(start_color='1F4E78', end_color='1F4E78', fill_type='solid')
                        cell.font = Font(bold=True, color='FFFFFF')
                    else:
                        cell.fill = PatternFill(start_color=cores_secoes['ambiente'], end_color=cores_secoes['ambiente'], fill_type='solid')
                    cell.border = Border(
                        left=Side(style='thin'),
                        right=Side(style='thin'),
                        top=Side(style='thin'),
                        bottom=Side(style='thin')
                    )
                    cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # 3. TOTAIS DIÁRIOS
            start_row = len(autor_df) + len(env_df) + 7
            resumo_sheet.merge_cells(f'A{start_row}:H{start_row}')
            titulo_cell = resumo_sheet[f'A{start_row}']
            titulo_cell.value = "TOTAIS DIÁRIOS"
            titulo_cell.font = titulo_font
            titulo_cell.alignment = Alignment(horizontal='center')
            titulo_cell.fill = PatternFill(start_color=cores_secoes['diario'], end_color=cores_secoes['diario'], fill_type='solid')
            
            # Adiciona os dados diários
            start_row += 1
            daily_totals = daily_df.groupby('data').sum().reset_index()
            for r_idx, row in enumerate(dataframe_to_rows(daily_totals, index=False, header=True), start_row):
                for c_idx, value in enumerate(row, 1):
                    cell = resumo_sheet.cell(row=r_idx, column=c_idx, value=value)
                    if r_idx == start_row:  # Cabeçalho
                        cell.font = Font(bold=True)
                        cell.fill = PatternFill(start_color='1F4E78', end_color='1F4E78', fill_type='solid')
                        cell.font = Font(bold=True, color='FFFFFF')
                    else:
                        cell.fill = PatternFill(start_color=cores_secoes['diario'], end_color=cores_secoes['diario'], fill_type='solid')
                    cell.border = Border(
                        left=Side(style='thin'),
                        right=Side(style='thin'),
                        top=Side(style='thin'),
                        bottom=Side(style='thin')
                    )
                    cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # 4. TOTAIS MENSAIS
            start_row = len(autor_df) + len(env_df) + len(daily_totals) + 10
            resumo_sheet.merge_cells(f'A{start_row}:H{start_row}')
            titulo_cell = resumo_sheet[f'A{start_row}']
            titulo_cell.value = "TOTAIS MENSAIS"
            titulo_cell.font = titulo_font
            titulo_cell.alignment = Alignment(horizontal='center')
            titulo_cell.fill = PatternFill(start_color=cores_secoes['mensal'], end_color=cores_secoes['mensal'], fill_type='solid')
            
            # Adiciona os dados mensais
            start_row += 1
            monthly_totals = monthly_df.groupby('mes').sum().reset_index()
            for r_idx, row in enumerate(dataframe_to_rows(monthly_totals, index=False, header=True), start_row):
                for c_idx, value in enumerate(row, 1):
                    cell = resumo_sheet.cell(row=r_idx, column=c_idx, value=value)
                    if r_idx == start_row:  # Cabeçalho
                        cell.font = Font(bold=True)
                        cell.fill = PatternFill(start_color='1F4E78', end_color='1F4E78', fill_type='solid')
                        cell.font = Font(bold=True, color='FFFFFF')
                    else:
                        cell.fill = PatternFill(start_color=cores_secoes['mensal'], end_color=cores_secoes['mensal'], fill_type='solid')
                    cell.border = Border(
                        left=Side(style='thin'),
                        right=Side(style='thin'),
                        top=Side(style='thin'),
                        bottom=Side(style='thin')
                    )
                    cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # Adiciona as outras abas
            if not daily_df.empty:
                daily_df.to_excel(writer, sheet_name='Estatísticas Diárias', index=False)
                self._format_sheet(writer.sheets['Estatísticas Diárias'])
            
            if not monthly_df.empty:
                monthly_df.to_excel(writer, sheet_name='Estatísticas Mensais', index=False)
                self._format_sheet(writer.sheets['Estatísticas Mensais'])
            
            if not branch_df.empty:
                branch_df.to_excel(writer, sheet_name='Estatísticas por Branch', index=False)
                self._format_sheet(writer.sheets['Estatísticas por Branch'])
            
            # Adiciona aba com informações dos autores
            authors_df = pd.DataFrame([
                {'Nome': author.name, 'Email': author.email}
                for author in self.authors
            ])
            authors_df.to_excel(writer, sheet_name='Informações', index=False)
            self._format_sheet(writer.sheets['Informações'])
            
            # Ajusta larguras das colunas em todas as abas
            for sheet in writer.sheets.values():
                for column in sheet.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                        
                    adjusted_width = (max_length + 2)
                    sheet.column_dimensions[column_letter].width = min(adjusted_width, 30)
        
        return str(excel_path)
    
    def _format_sheet(self, worksheet):
        """Aplica formatação padrão para uma aba."""
        # Cores
        header_fill = PatternFill(start_color='1F4E78', end_color='1F4E78', fill_type='solid')
        header_font = Font(color='FFFFFF', bold=True)
        
        # Borda
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Formata cabeçalho
        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border
        
        # Formata células de dados
        for row in worksheet.iter_rows(min_row=2):
            for cell in row:
                cell.border = border
                cell.alignment = Alignment(horizontal='center', vertical='center')
                
                # Aplica cores alternadas nas linhas
                if cell.row % 2 == 0:
                    cell.fill = PatternFill(start_color='F5F5F5', end_color='F5F5F5', fill_type='solid')
        
        # Adiciona filtros
        if worksheet.max_row > 1:
            worksheet.auto_filter.ref = worksheet.dimensions 